from tkinter import W
import warnings
from datetime import datetime

from numpy import sort
from exceptions import IncorrectLayout
import openpyxl as xl
import regex as re
from collections import defaultdict
import pandas as pd

from bom import Litze

# * get rid of the warnings for excel's not supported functionaltiy by openpyxl.
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


class Macro:
    def __init__(self, original_path=None):
        if original_path == None:
            self.original_path = "../../excel_files/test_file.xlsx"
        else:
            self.original_path = original_path
        self.new_file_path = "../../excel_files/output_files/macro_output.xlsx"
        self.wb = xl.load_workbook(self.original_path, data_only=True)
        self.dictionary_of_products: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
        self.litze_bom = Litze()

    def save_excel(self):
        self.wb.save(self.new_file_path)

    def put_number_to_new_sheet(self):
        df = pd.DataFrame(data=self.dictionary_of_products)
        # * to append ExcelWriter needs to be used.
        with pd.ExcelWriter(self.new_file_path, mode="a") as writer:
            df.to_excel(writer, sheet_name="outcome_macro")

    def open_ws(self):
        # * open workbook na Tabelle1 is existent, if not, just first sheet
        if "Tabelle1" in self.wb.sheetnames:
            return self.wb["Tabelle1"]
        return self.wb.active

    def clean_columns(self, ws: xl.Workbook):
        # * get rid of all irrelevant columns
        for row in ws.iter_rows(min_row=1, max_row=1):
            headers = ("Art", "Nr.", "Menge", "Beschreibung")
            for cell in row:
                if cell.value not in headers:
                    ws.delete_cols(cell.column)

    def clean_rows_from_art_ressource(self, ws: xl.Workbook):
        # * after _clean_columns all items with "Art" == system or Ressouce need to be removed
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value == "Art":
                    for col in ws.iter_cols(min_col=cell.column, max_col=cell.column):
                        for cell in col:
                            if cell.value == "System" or cell.value == "Ressource":
                                ws.delete_rows(cell.row)

    def remove_not_anchor_products(self, ws: xl.Workbook):
        # * consumption of components will be counted only for Anchors (** LT **** ****).
        # * blank spaces have to stay because description in "Beschreibung" of each Anchor is > one cell.
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value == "Nr.":
                    for col in ws.iter_cols(min_row=2, min_col=cell.column, max_col=cell.column):
                        for cell in col:
                            if cell.value:
                                m = re.search("\d{2}\sLT\s\d{4}\s\d{4}", str(cell.value))
                                if m is None:
                                    ws.delete_rows(cell.row)

    def check_columns_order(self, ws: xl.Workbook):
        # * checks if order of columns (after cleanup) is as intended.
        required_order = ("Art", "Nr.", "Menge", "Beschreibung")
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            if not row == required_order:
                raise IncorrectLayout

    def count_total_length(self, ws: xl.Workbook):  #!redundant
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value == "Nr.":
                    for column in ws.iter_cols(min_row=cell.row + 1, min_col=cell.column, max_col=cell.column):
                        for cell in column:
                            if (
                                cell.value not in self.dictionary_of_products.keys() and cell.value is not None
                            ):  # * empty cell = None.
                                self.dictionary_of_products[cell.value] = cell.offset(
                                    0, 1
                                ).value  # * layout article:quantity.
                            elif cell.value in self.dictionary_of_products.keys():
                                self.dictionary_of_products[cell.value] += cell.offset(0, 1).value

    def loop_ankers_sum_lengths(self, ws: xl.Workbook):
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value == "Nr.":
                    for column in ws.iter_cols(min_row=cell.row + 1, min_col=cell.column, max_col=cell.column):
                        for cell in column:
                            if cell.value is not None:
                                all_lenghts_str = self._get_all_lenghts_per_anker(cell)
                                all_lenghts_int = tuple(map(float, all_lenghts_str))
                                stk, gesamt, lv, lfr, uli, ure = all_lenghts_int
                                gesamt = stk * gesamt
                                lv = stk * lv
                                lfr = stk * lfr
                                uli = stk * uli
                                ure = stk * ure
                                self.dictionary_of_products[cell.value]["stk"] += stk
                                self.dictionary_of_products[cell.value]["gesamt"] += gesamt
                                self.dictionary_of_products[cell.value]["lv"] += lv
                                self.dictionary_of_products[cell.value]["lfr"] += lfr
                                self.dictionary_of_products[cell.value]["uli"] += uli
                                self.dictionary_of_products[cell.value]["ure"] += ure

    def _get_all_lenghts_per_anker(self, cell):
        stk, gesamt, lv = self._find_all_lengths_single_string(cell.offset(1, 2).value)
        lft, uli, ure = self._find_all_lengths_single_string(cell.offset(2, 2).value)
        all_lengths = [stk, gesamt, lv, lft, uli, ure]
        all_lengths_with_dots_decimal = [length.replace(",", ".") for length in all_lengths]
        return all_lengths_with_dots_decimal

    def _find_all_lengths_single_string(self, string: str):
        # * 48 Stk. à 19,5 m, Lv 6m
        # * Lfr 12,5m, Üli 1m, Üre 0m
        # pattern = re.compile(r"[a-zA-Z]*\s*(\d+,*\d*).*\s(\d+,*\d*).*(\d+,*\d*)")
        # match = pattern.findall(string)
        match = re.search("\D*\s*(\d+,*\d*).*\s(\d+,*\d*).*\s(\d+,*\d*)", string)
        # * returns a tuple of strings
        return match.groups()

    def calculate_consumption_loop(self):
        for key in self.dictionary_of_products.keys():
            # *mutlipliers from BOM
            steel_multiplier = self.litze_bom.material_multipliers_strand[key[:2]]["steel"]
            pe_multiplier = self.litze_bom.material_multipliers_strand[key[:2]]["PE"]
            gfk_multiplier = self.litze_bom.material_multipliers_strand[key[:2]]["GFK"]
            schrumpfschlauch_multiplier = self.litze_bom.material_multipliers_strand[key[:2]]["schrumpfschlauch"]
            azban_mutliplier = self.litze_bom.material_multipliers_strand[key[:2]]["AZBAN"]
            # * lengths for each type.
            stk = self.dictionary_of_products[key]["stk"]
            gesamt = self.dictionary_of_products[key]["gesamt"]
            lfr = self.dictionary_of_products[key]["lfr"]

            L00VLI01, PE, GFK, Schrumpfschlauch, AZBAN5 = self.litze_bom.calculate_consumption(
                gesamt,
                lfr,
                stk,
                steel_multiplier,
                pe_multiplier,
                gfk_multiplier,
                schrumpfschlauch_multiplier,
                azban_mutliplier,
            )
            self._format_consumption(key, L00VLI01, PE, GFK, Schrumpfschlauch, AZBAN5)

    def _format_consumption(self, key, L00VLI01, PE, GFK, Schrumpfschlauch, AZBAN5):
        self.dictionary_of_products[key]["L00VLI01"] = str(round(L00VLI01, 4)) + " KG"
        self.dictionary_of_products[key]["PE"] = str(round(PE, 4)) + " MET"
        self.dictionary_of_products[key]["GFK"] = str(round(GFK, 4)) + " ROL"
        self.dictionary_of_products[key]["Schrumpfschlauch"] = str(round(Schrumpfschlauch, 4)) + " MET"
        self.dictionary_of_products[key]["AZBAN5"] = str(round(AZBAN5, 4)) + " ROL"
